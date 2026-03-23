import io
from typing import List, Any, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_sheet(ws, title: str, headers: List[str], rows: List[List[Any]]) -> None:
    ws.title = title

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-width
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows) + 2)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 22


def build_workbook(sheets: List[Dict]) -> bytes:
    """
    sheets: [{"title": str, "headers": [...], "rows": [[...], ...]}, ...]
    Returns xlsx bytes.
    """
    wb = Workbook()
    for i, sheet in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        _write_sheet(ws, sheet["title"], sheet["headers"], sheet["rows"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
